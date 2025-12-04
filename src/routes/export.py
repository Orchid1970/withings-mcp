"""
Export routes for Withings data.

This module provides endpoints to export Withings health data to Excel format.
Standalone script reference: src/export_withings_to_excel.py
"""

import io
import logging
from datetime import datetime, timedelta
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
import httpx
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["export"])

# Styling constants (Carrot Orange, Deep Blue, Grey palette)
CARROT_ORANGE = "FF9933"
DEEP_BLUE = "003366"
GREY_LIGHT = "F5F5F5"
GREY_DARK = "CCCCCC"
WHITE = "FFFFFF"

# Timezone for timestamps
TIMEZONE = ZoneInfo("America/Los_Angeles")


def get_withings_data(base_url: str, token: str, days: int) -> dict:
    """
    Fetch all Withings data from the withings-mcp service.
    
    Args:
        base_url: Base URL of withings-mcp service (e.g., https://withings-mcp-production.up.railway.app)
        token: Withings access token
        days: Number of days of historical data to fetch
    
    Returns:
        Dictionary containing all Withings data
    """
    try:
        headers = {"Authorization": f"Bearer {token}"}
        
        # Fetch all data from the service
        response = httpx.get(
            f"{base_url}/data/all",
            headers=headers,
            timeout=30.0
        )
        response.raise_for_status()
        
        data = response.json()
        logger.info(f"Successfully fetched Withings data from {base_url}")
        return data
        
    except httpx.HTTPError as e:
        logger.error(f"Failed to fetch Withings data: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch Withings data: {str(e)}")


def create_excel_workbook(data: dict) -> bytes:
    """
    Create an Excel workbook from Withings data.
    
    Args:
        data: Dictionary containing Withings data
    
    Returns:
        Bytes of the Excel workbook
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_fill = PatternFill(start_color=CARROT_ORANGE, end_color=CARROT_ORANGE, fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color=WHITE)
    
    alt_fill = PatternFill(start_color=GREY_LIGHT, end_color=GREY_LIGHT, fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Sheet 1: Activity Summary
    ws_activity = wb.create_sheet("Activity_Summary")
    activity_data = data.get("activity", [])
    
    if activity_data:
        headers = ["Date", "Steps", "Distance (m)", "Calories", "Active Time (min)"]
        ws_activity.append(headers)
        
        for cell in ws_activity[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for idx, record in enumerate(activity_data[:30], start=2):  # Last 30 days
            date_str = record.get("date", "N/A")
            steps = record.get("steps", 0)
            distance = record.get("distance", 0)
            calories = record.get("calories", 0)
            active_time = record.get("active_time", 0)
            
            ws_activity.append([date_str, steps, distance, calories, active_time])
            
            if idx % 2 == 0:
                for cell in ws_activity[idx]:
                    cell.fill = alt_fill
                    cell.border = border
        
        ws_activity.column_dimensions["A"].width = 15
        ws_activity.column_dimensions["B"].width = 12
        ws_activity.column_dimensions["C"].width = 15
        ws_activity.column_dimensions["D"].width = 12
        ws_activity.column_dimensions["E"].width = 18
    
    # Sheet 2: Body Composition (Latest)
    ws_body = wb.create_sheet("Body_Composition_Latest")
    metrics = data.get("metrics", [])
    
    if metrics:
        headers = ["Metric Type", "Value", "Unit", "Timestamp"]
        ws_body.append(headers)
        
        for cell in ws_body[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for idx, metric in enumerate(metrics[:20], start=2):
            metric_type = metric.get("type", "N/A")
            value = metric.get("value", "N/A")
            unit = metric.get("unit", "N/A")
            timestamp = metric.get("timestamp", "N/A")
            
            ws_body.append([metric_type, value, unit, timestamp])
            
            if idx % 2 == 0:
                for cell in ws_body[idx]:
                    cell.fill = alt_fill
                    cell.border = border
        
        ws_body.column_dimensions["A"].width = 20
        ws_body.column_dimensions["B"].width = 15
        ws_body.column_dimensions["C"].width = 12
        ws_body.column_dimensions["D"].width = 25
    
    # Sheet 3: Blood Pressure (Last Readings)
    ws_bp = wb.create_sheet("Last_Readings_All")
    blood_pressure = data.get("blood_pressure", [])
    
    if blood_pressure:
        headers = ["Timestamp", "Systolic (mmHg)", "Diastolic (mmHg)", "Heart Rate (bpm)"]
        ws_bp.append(headers)
        
        for cell in ws_bp[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for idx, reading in enumerate(blood_pressure[:20], start=2):
            timestamp = reading.get("timestamp", "N/A")
            systolic = reading.get("systolic", "N/A")
            diastolic = reading.get("diastolic", "N/A")
            heart_rate = reading.get("heart_rate", "N/A")
            
            ws_bp.append([timestamp, systolic, diastolic, heart_rate])
            
            if idx % 2 == 0:
                for cell in ws_bp[idx]:
                    cell.fill = alt_fill
                    cell.border = border
        
        ws_bp.column_dimensions["A"].width = 25
        ws_bp.column_dimensions["B"].width = 18
        ws_bp.column_dimensions["C"].width = 18
        ws_bp.column_dimensions["D"].width = 18
    
    # Sheet 4: Export Metadata
    ws_meta = wb.create_sheet("Export_Metadata")
    now = datetime.now(TIMEZONE).isoformat()
    ws_meta.append(["Export Timestamp", now])
    ws_meta.append(["Data Source", "Withings API"])
    ws_meta.append(["Service", "withings-mcp"])
    ws_meta.append(["Timezone", "America/Los_Angeles"])
    
    for row in ws_meta.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = border
    
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 40
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/excel")
async def export_excel(
    days: int = Query(7, ge=1, le=365, description="Number of days of data to export"),
    token: Optional[str] = Query(None, description="Withings access token (optional, uses env var if not provided)")
) -> StreamingResponse:
    """
    Export Withings data to Excel workbook.
    
    Query Parameters:
    - days: Number of days of historical data (default: 7, max: 365)
    - token: Withings access token (optional, uses WITHINGS_ACCESS_TOKEN env var if not provided)
    
    Returns:
        Excel workbook as downloadable attachment
    
    Example:
        GET /export/excel?days=30
        GET /export/excel?days=7&token=YOUR_TOKEN
    
    Note:
        This endpoint uses the standalone script logic from src/export_withings_to_excel.py
        For local usage, see that file for the Python script version.
    """
    try:
        # Get token from parameter or environment
        access_token = token or os.getenv("WITHINGS_ACCESS_TOKEN")
        if not access_token:
            raise HTTPException(
                status_code=400,
                detail="Withings access token required. Provide via ?token=YOUR_TOKEN or set WITHINGS_ACCESS_TOKEN env var"
            )
        
        # Get base URL
        base_url = os.getenv("WITHINGS_MCP_BASE_URL", "https://withings-mcp-production.up.railway.app")
        
        # Fetch data
        logger.info(f"Fetching Withings data for last {days} days")
        data = get_withings_data(base_url, access_token, days)
        
        # Create workbook
        logger.info("Creating Excel workbook")
        excel_bytes = create_excel_workbook(data)
        
        # Return as downloadable file
        now = datetime.now(TIMEZONE).strftime("%Y%m%d_%H%M%S")
        filename = f"withings_export_{now}.xlsx"
        
        return StreamingResponse(
            iter([excel_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")


@router.post("/excel")
async def export_excel_post(
    days: int = Query(7, ge=1, le=365, description="Number of days of data to export"),
    token: Optional[str] = Query(None, description="Withings access token (optional, uses env var if not provided)")
) -> StreamingResponse:
    """
    Export Withings data to Excel workbook (POST method).
    
    Same as GET /export/excel but accepts POST requests.
    """
    return await export_excel(days=days, token=token)
